"""
输出文件生成模块
用于生成文库表和芯片表Excel文件
"""
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
import openpyxl


class OutputGenerator:
    """输出文件生成器"""
    
    def __init__(self, output_dir: str = "output"):
        """
        初始化输出生成器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    @staticmethod
    def _read_headers_from_template(template_file: str, preferred_sheet: Optional[str] = None) -> Optional[List[str]]:
        try:
            wb = openpyxl.load_workbook(template_file, data_only=True)
            ws = None
            if preferred_sheet and preferred_sheet in wb.sheetnames:
                ws = wb[preferred_sheet]
            else:
                ws = wb[wb.sheetnames[0]]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() for h in headers if h is not None and str(h).strip() != ""]
            return headers if headers else None
        except Exception:
            return None

    @staticmethod
    def _write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]]):
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    
    def generate_library_table(self, libraries: List[Dict[str, Any]], 
                               template_file: Optional[str] = None,
                               output_file: Optional[str] = None) -> str:
        """
        生成文库表
        
        Args:
            libraries: 文库列表
            template_file: 模板文件路径（可选）
            output_file: 输出文件路径（可选）
            
        Returns:
            输出文件路径
        """
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = self.output_dir / f"文库表_{timestamp}.xlsx"
        else:
            output_file = Path(output_file)
        
        columns = None
        if template_file and Path(template_file).exists():
            columns = self._read_headers_from_template(template_file, preferred_sheet="文库表模版")

        if not columns:
            # 默认列（尽量覆盖模板/示例的常用字段）
            columns = [
                "芯片",
                "芯片数据量",
                "上机时间",
                "分析时间",
                "样本名称",
                "文库编号",
                "index",
                "Clean Reads",
                "≥Q20%",
                "Q30",
                "物种名称",
                "分类",
                "taxid",
                "拉丁文",
                "内部对照spike.1RPM值",
                "rpm",
                "uniq rpm",
            ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "文库表"
        self._write_sheet(ws, columns, libraries)
        wb.save(output_file)
        
        return str(output_file)
    
    def generate_chip_table(self, chips: List[Dict[str, Any]],
                           template_file: Optional[str] = None,
                           output_file: Optional[str] = None) -> str:
        """
        生成芯片表
        
        Args:
            chips: 芯片列表
            template_file: 模板文件路径（可选）
            output_file: 输出文件路径（可选）
            
        Returns:
            输出文件路径
        """
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = self.output_dir / f"芯片表_{timestamp}.xlsx"
        else:
            output_file = Path(output_file)
        
        columns = None
        if template_file and Path(template_file).exists():
            columns = self._read_headers_from_template(template_file, preferred_sheet="芯片表模版")

        if not columns:
            columns = ["实验项目", "测序日期", "测序仪SN", "Run数", "芯片SN", "测序仪型号", "试验结果", "备注2"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "芯片表"
        self._write_sheet(ws, columns, chips)
        wb.save(output_file)
        
        return str(output_file)
    
    def generate_combined_output(self, libraries: List[Dict[str, Any]],
                                 chips: List[Dict[str, Any]],
                                 output_file: Optional[str] = None,
                                 lib_template_file: Optional[str] = None,
                                 chip_template_file: Optional[str] = None) -> str:
        """
        生成合并输出文件（包含文库表和芯片表）
        
        Args:
            libraries: 文库列表
            chips: 芯片列表
            output_file: 输出文件路径（可选）
            
        Returns:
            输出文件路径
        """
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = self.output_dir / f"排布结果_{timestamp}.xlsx"
        else:
            output_file = Path(output_file)

        # 严格按模板列顺序（如果提供），不改变模板列排布
        lib_columns = None
        chip_columns = None
        if lib_template_file and Path(lib_template_file).exists():
            lib_columns = self._read_headers_from_template(lib_template_file, preferred_sheet="文库表模版")
        if chip_template_file and Path(chip_template_file).exists():
            chip_columns = self._read_headers_from_template(chip_template_file, preferred_sheet="芯片表模版")

        if not lib_columns:
            lib_columns = list(libraries[0].keys()) if libraries else ["芯片", "样本名称", "文库编号", "index", "物种名称"]
        if not chip_columns:
            chip_columns = list(chips[0].keys()) if chips else ["实验项目", "测序日期", "测序仪SN", "Run数", "芯片SN"]

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "芯片表"
        self._write_sheet(ws1, chip_columns, chips)

        ws2 = wb.create_sheet("文库表")
        self._write_sheet(ws2, lib_columns, libraries)

        wb.save(output_file)
        return str(output_file)
