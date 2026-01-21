"""
输入文件解析模块
用于解析输入表、规则文件、对照文件等
"""
from typing import Dict, List, Any, Optional
from pathlib import Path
import re

import openpyxl


class InputParser:
    """输入表解析器"""
    
    def __init__(self, input_file: str):
        """
        初始化输入表解析器
        
        Args:
            input_file: 输入表Excel文件路径
        """
        self.input_file = Path(input_file)
        self.wb = None
        
    def parse(self) -> Dict[str, Any]:
        """
        解析输入表
        
        Returns:
            包含meta和samples的字典
        """
        if not self.input_file.exists():
            raise FileNotFoundError(f"输入表不存在: {self.input_file}")

        self.wb = openpyxl.load_workbook(self.input_file, data_only=True)
        ws = self.wb[self.wb.sheetnames[0]]

        meta: Dict[str, Any] = {}
        samples: List[Dict[str, Any]] = []

        # 输入表是“配置项(key,value,备注)”+“样本行(样本名,物种)”的结构
        sample_id_re = re.compile(r"^[A-Za-z]-\d{4}-\d{2}$")  # 例如 F-0020-01

        def split_semicolon(val: Any) -> List[str]:
            """
            支持英文分号;（也兼容中文；）分隔，去掉空项并strip。
            """
            if val is None:
                return []
            s = str(val).strip()
            if s == "":
                return []
            s = s.replace("；", ";")
            parts = [p.strip() for p in s.split(";")]
            return [p for p in parts if p != ""]

        for r in range(1, ws.max_row + 1):
            k = ws.cell(r, 1).value
            v1 = ws.cell(r, 2).value
            v2 = ws.cell(r, 3).value
            v3 = ws.cell(r, 4).value
            if k is None or str(k).strip() == "":
                continue

            k_str = str(k).strip()

            # 样本行：A列为样本名，B列为病原体/物种名称
            if sample_id_re.match(k_str):
                pathogens = split_semicolon(v1)
                rpm_ranges = split_semicolon(v2)
                spike_ranges = split_semicolon(v3)

                # 配对规则：
                # - Value1: 病原体列表（必需）
                # - Value2: rpm范围列表；若只给1个则应用到所有病原体；若给N个则需与病原体数量一致
                # - Value3: spike-rpm范围列表；同上
                species: List[Dict[str, Any]] = []
                n = len(pathogens)
                if n > 0:
                    for i, name in enumerate(pathogens):
                        rpm_i = ""
                        spike_i = ""
                        if len(rpm_ranges) == 1:
                            rpm_i = rpm_ranges[0]
                        elif len(rpm_ranges) == n:
                            rpm_i = rpm_ranges[i]
                        elif len(rpm_ranges) == 0:
                            rpm_i = ""
                        else:
                            # 数量不一致时：保留空，避免错误阻断（也可改成 raise）
                            rpm_i = ""

                        if len(spike_ranges) == 1:
                            spike_i = spike_ranges[0]
                        elif len(spike_ranges) == n:
                            spike_i = spike_ranges[i]
                        elif len(spike_ranges) == 0:
                            spike_i = ""
                        else:
                            spike_i = ""

                        species.append({"name": name, "rpm_range": rpm_i, "spike_rpm_range": spike_i})

                samples.append({"sample_id": k_str, "species": species})
                continue

            # 配置项
            meta[k_str] = v1
            # 保存多维值（目前除病原体外一般只有value1）
            meta[f"{k_str}__value2"] = v2
            meta[f"{k_str}__value3"] = v3

        # 解析测序仪信息（按“测序仪1-SN / 测序仪1-RUN ...”）
        sequencers: List[Dict[str, Any]] = []
        try:
            n = int(meta.get("需要用到的测序仪台数") or 0)
        except Exception:
            n = 0
        for i in range(1, n + 1):
            sn = meta.get(f"测序仪{i}-SN")
            run = meta.get(f"测序仪{i}-RUN")
            if sn is None or str(sn).strip() == "":
                continue
            sequencers.append(
                {
                    "index": i,
                    "sn": str(sn).strip(),
                    "run": "" if run is None else str(run).strip(),
                }
            )

        meta["sequencers"] = sequencers
        return {"meta": meta, "samples": samples}


class RulesParser:
    """规则文件解析器"""
    
    def __init__(self, rules_file: str):
        """
        初始化规则解析器
        
        Args:
            rules_file: 规则Excel文件路径
        """
        self.rules_file = Path(rules_file)
        self.rules = {}
        
    def parse(self) -> Dict[str, Any]:
        """
        解析规则文件
        
        Returns:
            规则字典
        """
        if not self.rules_file.exists():
            return {}
        try:
            wb = openpyxl.load_workbook(self.rules_file, data_only=True)
            ws = wb[wb.sheetnames[0]]

            # 规则表格式：A列为字段名，B列为规则描述，C列为示例
            for r in range(1, ws.max_row + 1):
                k = ws.cell(r, 1).value
                if k is None or str(k).strip() == "":
                    continue
                key = str(k).strip()
                self.rules[key] = {
                    "rule": ws.cell(r, 2).value,
                    "example": ws.cell(r, 3).value,
                }
            return self.rules
        except Exception:
            return {}


class ReferenceParser:
    """参考文件解析器（NC、PC、物种列表等）"""
    
    def __init__(self, nc_file: Optional[str] = None, 
                 pc_file: Optional[str] = None,
                 species_file: Optional[str] = None,
                 sequencer_file: Optional[str] = None):
        """
        初始化参考文件解析器
        
        Args:
            nc_file: 阴性对照文件路径
            pc_file: 阳性对照文件路径
            species_file: 物种列表文件路径
            sequencer_file: 测序仪对应关系文件路径
        """
        self.nc_file = Path(nc_file) if nc_file else None
        self.pc_file = Path(pc_file) if pc_file else None
        self.species_file = Path(species_file) if species_file else None
        self.sequencer_file = Path(sequencer_file) if sequencer_file else None
        
    def parse_nc(self) -> List[Dict[str, Any]]:
        """解析阴性对照文件"""
        if not self.nc_file or not self.nc_file.exists():
            return []
        try:
            wb = openpyxl.load_workbook(self.nc_file, data_only=True)
            # 优先找“NC列表”
            ws = wb["NC列表"] if "NC列表" in wb.sheetnames else wb[wb.sheetnames[0]]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() for h in headers if h is not None and str(h).strip() != ""]
            if not headers:
                return []
            out: List[Dict[str, Any]] = []
            for r in range(2, ws.max_row + 1):
                row = {headers[i - 1]: ws.cell(r, i).value for i in range(1, len(headers) + 1)}
                # 跳过全空行
                if all(v in (None, "") for v in row.values()):
                    continue
                out.append(row)
            return out
        except Exception:
            return []
    
    def parse_pc(self) -> List[Dict[str, Any]]:
        """解析阳性对照文件"""
        if not self.pc_file or not self.pc_file.exists():
            return []
        try:
            wb = openpyxl.load_workbook(self.pc_file, data_only=True)
            # 优先找“PC列表”
            ws = wb["PC列表"] if "PC列表" in wb.sheetnames else wb[wb.sheetnames[0]]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            headers = [str(h).strip() for h in headers if h is not None and str(h).strip() != ""]
            if not headers:
                return []
            out: List[Dict[str, Any]] = []
            for r in range(2, ws.max_row + 1):
                row = {headers[i - 1]: ws.cell(r, i).value for i in range(1, len(headers) + 1)}
                if all(v in (None, "") for v in row.values()):
                    continue
                out.append(row)
            return out
        except Exception:
            return []
    
    def parse_species(self) -> Dict[str, Any]:
        """解析物种列表文件"""
        if not self.species_file or not self.species_file.exists():
            return {}
        try:
            wb = openpyxl.load_workbook(self.species_file, data_only=True)
            # 优先找“物种列表/物种”sheet
            ws = None
            for cand in ["物种列表", "物种"]:
                if cand in wb.sheetnames:
                    ws = wb[cand]
                    break
            if ws is None:
                ws = wb[wb.sheetnames[0]]

            # 找表头行
            header_row = None
            headers: List[str] = []
            for r in range(1, min(ws.max_row, 20) + 1):
                row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 30) + 1)]
                if any(v == "名称" for v in row) and any(v == "taxid" for v in row):
                    header_row = r
                    headers = ["" if v is None else str(v).strip() for v in row]
                    break
            if header_row is None:
                return {}

            # 建映射：名称 -> {分类,taxid,拉丁文...}
            name_idx = headers.index("名称") + 1
            result: Dict[str, Any] = {}
            for r in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(r, name_idx).value
                if name is None or str(name).strip() == "":
                    continue
                rec: Dict[str, Any] = {}
                for c, h in enumerate(headers, start=1):
                    if h == "":
                        continue
                    rec[h] = ws.cell(r, c).value
                result[str(name).strip()] = rec
            return result
        except Exception:
            return {}
    
    def parse_sequencer(self) -> Dict[str, Any]:
        """解析测序仪对应关系文件"""
        if not self.sequencer_file or not self.sequencer_file.exists():
            return {}
        try:
            wb = openpyxl.load_workbook(self.sequencer_file, data_only=True)
            ws = wb[wb.sheetnames[0]]

            # 表头：序号,设备名称,内部编号,设备型号,设备序列号
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            # 找关键列
            try:
                model_col = headers.index("设备型号") + 1
                sn_col = headers.index("设备序列号") + 1
            except Exception:
                return {}

            mapping: Dict[str, Any] = {}
            for r in range(2, ws.max_row + 1):
                sn = ws.cell(r, sn_col).value
                model = ws.cell(r, model_col).value
                if sn is None or str(sn).strip() == "":
                    continue
                mapping[str(sn).strip()] = {"设备型号": model}
            return mapping
        except Exception:
            return {}
