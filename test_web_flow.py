#!/usr/bin/env python3
"""
端到端测试：模拟 Web 前端三步流程
1) /api/generate-chips
2) /api/generate-libraries
3) /api/generate-files

输入使用 attachments/Input table.xlsx 以及默认 attachments/*.xlsx
"""

from __future__ import annotations

from pathlib import Path
import json
import re

import openpyxl

from app import app


def _post_multipart(client, url: str, files: dict, form: dict | None = None):
    data = {}
    if form:
        data.update(form)
    for field, path in files.items():
        p = Path(path)
        data[field] = (p.open("rb"), p.name)
    return client.post(url, data=data, content_type="multipart/form-data")


def main():
    root = Path(__file__).parent
    input_xlsx = root / "attachments" / "Input table.xlsx"
    assert input_xlsx.exists(), f"missing {input_xlsx}"

    # 1) generate chips
    with app.test_client() as client:
        resp = _post_multipart(
            client,
            "/api/generate-chips",
            files={
                "input_file": str(input_xlsx),
                # 不传其它文件则走默认 attachments/*
            },
            form={},
        )
        assert resp.status_code == 200, resp.data
        j = resp.get_json()
        assert j["success"] is True, j
        chips = j["chips"]
        assert len(chips) > 0

        # run increment per sequencer
        last_run = {}
        for c in chips:
            sn = c["测序仪SN"]
            run = int(c["Run数"]) if str(c["Run数"]).strip() != "" else 0
            if sn in last_run:
                assert run == last_run[sn] + 1, (sn, last_run[sn], run)
            last_run[sn] = run

        # 2) generate libraries preview
        resp2 = client.post(
            "/api/generate-libraries",
            data=json.dumps({"chips": chips}),
            content_type="application/json",
        )
        assert resp2.status_code == 200, resp2.data
        j2 = resp2.get_json()
        assert j2["success"] is True, j2
        libs = j2["libraries"]
        assert len(libs) > 0

        # PC/NC present
        pc_rows = [r for r in libs if str(r.get("样本名称", "")).endswith("CN-PC")]
        nc_rows = [r for r in libs if str(r.get("样本名称", "")).endswith("CN-NC")]
        assert len(pc_rows) > 0, "PC rows missing"
        assert len(nc_rows) > 0, "NC rows missing"

        # PC rows should have rpm + spike-rpm filled
        pc_by_chip = {}
        for r in pc_rows:
            pc_by_chip.setdefault(r["芯片"], []).append(r)
        first_chip, grp = next(iter(pc_by_chip.items()))
        assert grp[0].get("内部对照spike.1RPM值") not in (None, "", " "), grp[0]
        assert grp[0].get("rpm") not in (None, "", " "), grp[0]

        # PC rows share same index and lib id WITHIN SAME CHIP
        # pick first chip group
        pc_index = grp[0]["index"]
        pc_lib = grp[0]["文库编号"]
        assert all(r["index"] == pc_index for r in grp)
        assert all(r["文库编号"] == pc_lib for r in grp)

        # ensure no 上机时间.1 key (we removed)
        assert "上机时间.1" not in libs[0]

        # 3) generate files
        resp3 = client.post(
            "/api/generate-files",
            data=json.dumps({"chips": chips, "libraries": libs}),
            content_type="application/json",
        )
        assert resp3.status_code == 200, resp3.data
        j3 = resp3.get_json()
        assert j3["success"] is True, j3

        # download combined and inspect headers (must follow ref templates)
        dl = client.get("/api/download/combined")
        assert dl.status_code == 200, dl.data[:200]
        # 保存到临时文件并用 openpyxl 读取
        out_path = root / "test_output" / "web_combined_latest.xlsx"
        out_path.parent.mkdir(exist_ok=True)
        out_path.write_bytes(dl.data)

        wb = openpyxl.load_workbook(out_path, data_only=True)
        assert "文库表" in wb.sheetnames
        ws = wb["文库表"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # should match ref/文库表模版.xlsx 的列顺序（含重复的“上机时间”）
        tpl = openpyxl.load_workbook(root / "ref" / "文库表模版.xlsx", data_only=True)
        tpl_ws = tpl["文库表模版"]
        tpl_headers = [tpl_ws.cell(1, c).value for c in range(1, tpl_ws.max_column + 1)]
        while tpl_headers and (tpl_headers[-1] is None or str(tpl_headers[-1]).strip() == ""):
            tpl_headers.pop()
        assert headers == tpl_headers, (headers, tpl_headers)

        # spot-check that PC/NC appear in output sheet
        # find sample column index
        try:
            sample_col = headers.index("样本名称") + 1
        except ValueError:
            sample_col = None
        assert sample_col is not None
        values = set()
        for r in range(2, min(ws.max_row, 500) + 1):
            v = ws.cell(r, sample_col).value
            if v:
                values.add(str(v))
        assert any(v.endswith("CN-PC") for v in values), values
        assert any(v.endswith("CN-NC") for v in values), values

        # 接头号应全局连续（跨芯片不重置）：第二张芯片不应重新从A01开始
        # 只做弱校验：如果存在多张芯片，检查第二张芯片的第一条“普通样本”index不为A01
        # 这里用“样本名称形如 F-0000-00”过滤普通样本
        import re as _re
        sample_pat = _re.compile(r"^[A-Za-z]-\\d{4}-\\d{2}$")
        # 收集每条记录的(芯片,index,样本名称)
        lib_sheet = wb["文库表"]
        headers2 = headers
        chip_col = headers2.index("芯片") + 1 if "芯片" in headers2 else None
        idx_col = headers2.index("index") + 1 if "index" in headers2 else None
        sample_col = headers2.index("样本名称") + 1 if "样本名称" in headers2 else None
        if chip_col and idx_col and sample_col:
            first_by_chip = {}
            for r in range(2, lib_sheet.max_row + 1):
                chip_v = lib_sheet.cell(r, chip_col).value
                idx_v = lib_sheet.cell(r, idx_col).value
                s_v = lib_sheet.cell(r, sample_col).value
                if not chip_v or not idx_v or not s_v:
                    continue
                s_v = str(s_v)
                if not sample_pat.match(s_v):
                    continue
                first_by_chip.setdefault(str(chip_v), str(idx_v))
            chips_seen = list(first_by_chip.keys())
            if len(chips_seen) >= 2:
                assert first_by_chip[chips_seen[1]] != "A01", first_by_chip

    print("E2E WEB FLOW TEST: OK")


if __name__ == "__main__":
    main()

